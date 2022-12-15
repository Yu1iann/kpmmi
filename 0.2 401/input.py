import openpyxl

def input():
    wb = openpyxl.load_workbook('TP1.xlsx')
    #print(wb.sheetnames)
    sheet = wb['Лист1']
    #rows = sheet.max_row
    #cols = sheet.max_column
    #print(rows,cols)

    mr=[]
    nq=[]
    tq=[]
    TT=[]
    ll=sheet.cell(row=1, column=2).value
    cols=ll
    rows=ll
    print(ll)

    for i in range(3, rows + 3):
        rowi=[]
        for j in range(2, cols + 2):
            rowi.append( sheet.cell(row = i, column = j).value   )
        mr.append( rowi   )

        print(i, ' ',j,' ',rowi)

    for j in range(3, cols + 3):
        tq.append( sheet.cell(row = j, column = ll+5).value   )
        nq.append( sheet.cell(row = j, column = ll+4).value   )
    t1=[]
    t2=[]
    for j in range(3, cols + 3):
        t1.append(j-2)
        t2.append(sheet.cell(row = j, column = 12).value)

    TT.append(t1)
    TT.append(t2)
    del t1,t2
    print('nq ',nq)
    print('tq ',tq)
    print('TT ',TT)
    return nq, tq, TT,mr

def input1(rowpast):
    print('_______________input1___________________')
    wb = openpyxl.load_workbook('Расстояния.xlsx')
    #print(wb.sheetnames)
    sheet = wb['План_посещений']
    #rows = sheet.max_row
    #cols = sheet.max_column
    #print(rows,cols)
    mr=[]
    nq=[]
    tq=[]
    t1=[]
    t2=[]
    t3=[]
    t5=[]
    t4 = []
    TT=[]
    PARAM=  [[] for x in range(4)]
    TP0 = sheet.cell(row=rowpast+2, column=11).value
    ni=1
    row=rowpast+2
    while True:
        row += 1
        t2.append(sheet.cell(row=row - 1, column=1).value) # Код ТТ
        t3.append(sheet.cell(row=row - 1, column=3).value) # Название/адрес
        t4.append(sheet.cell(row=row - 1, column=6).value) # Адрес клиента
        t5.append(sheet.cell(row=row - 1, column=9).value)  # Количество посещений
        nq.append(sheet.cell(row=row - 1, column=9).value) # Количество посещений
        tq.append(sheet.cell(row=row - 1, column=12).value) # Продолжительность посещения
        TPi = sheet.cell(row=row, column=11).value
        if TP0==TPi:
            ni += 1
        else:break
    for i in range(ni):
        t1.append(i+1)
    TT.append(t1)
    TT.append(t2)
    print('_______________TT___________________')
    print (TT)
    print(PARAM)
    PARAM[0].extend(t2)
    #PARAM.append(t2)
    #PARAM[0]=t2[:]
    PARAM[1].extend(t3)
    PARAM[2].extend(t4)
    PARAM[3].extend(t5)
    print('_______________PARAM___________________')
    for i in range(4):
        print(PARAM[i])
    print (PARAM)
    rows=ni
    cols=ni

    sheetmr = wb['матрица расстояний']
    for i in range(rowpast,rowpast+rows):
        rowi=[]
        for j in range(rowpast,rowpast+cols):
            rowi.append( sheetmr.cell(row = i, column = j).value   )
        mr.append( rowi   )
    print('_______________MR___________________')
    for i in range(len(mr[0])):
        print(mr[i]) #матрица расстояний


        #print(i, ' ',j,' ',rowi)
    return TP0,PARAM,nq, tq, TT,mr